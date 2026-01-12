import openpyxl as xl                          #interacting with excel files
import numpy as np
from tkinter import Tk                         #gui for grabbing excel file
from tkinter.filedialog import askopenfilename 
import matplotlib.pyplot as plt                #for plotting graphs
import seaborn as sns                          #heatmap? not sure

Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
filename = askopenfilename() # show an "Open" dialog box and return the path to the selected file

wb = xl.load_workbook(filename)

sn = wb.sheetnames #vector containing sheetnames from xl file

nsb_values = np.zeros((6, 6)) #creating matrices that will contain the data to be graphed
stn_values = np.zeros((6, 6))

# Command line dialog for choosing which sheets in the file to use.
print("Sheet names in ", filename, ":", sep="")

for x in range(1, len(sn)+1):
    print(x,'. ',sn[x-1], sep="")

nsb_i = input("Enter the index for the sheet name containing the NSB data: ")

nsb_ws = wb[wb.sheetnames[int(nsb_i)-1]]

for i in range(2,8):
    for j in range(2,8):
        nsb_values[i-2, j-2] = nsb_ws.cell(row=i,column=j).value #Loads data matrices with data!


stn_i = input("Enter the index for the sheet name containing the signal to noise data: ")

stn_ws = wb[wb.sheetnames[int(stn_i)-1]]

for i in range(2,8):
    for j in range(2,8):
        stn_values[i-2, j-2] = stn_ws.cell(row=i,column=j).value

data = nsb_values
data_second = stn_values

# Normalize the data to scale from 0 to 1
normalized_data = (data - data.min()) / (data.max() - data.min())

# Plotting the heatmap
plt.figure(figsize=(8, 6))
ax = sns.heatmap(
    normalized_data,                # Normalized values from the first data set for coloring
    annot=data_second,              # Display values from the second data set
    fmt=".3f",                      # Format for text annotation
    cmap="Greys",                   # Grayscale color map
    cbar_kws={"label": "Non-specific binding/ 0 analyte TL/CL", "ticks": [0, 1]},  # Color bar for normalized range
    annot_kws={"weight": "bold"}    # Default bold text
)

# Adjust color bar ticks to match the actual data range
colorbar = ax.collections[0].colorbar
colorbar.set_ticks([0, 1])
colorbar.set_ticklabels([0, f"{data.max():.2f}"])  # Actual data range for color bar

# Adjust text color based on the normalized shade value for readability
for i, text in enumerate(ax.texts):
    row, col = divmod(i, data_second.shape[1])
    normalized_value = normalized_data[row, col]
    text.set_color("black" if normalized_value < 0.5 else "white")

# Labels and title
ax.set_title("IL6 antibody screening")
ax.set_yticks([0.5, 1.5, 2.5, 3.5, 4.5, 5.5], labels=['pAb-1', 'mAb-1', 'mAb-2', 'mAb-3', 'mAb-4', 'mAb-5'], rotation=0)
ax.set_xticks([0.5, 1.5, 2.5, 3.5, 4.5, 5.5], labels=['pAb-1', 'mAb-1', 'mAb-2', 'mAb-3', 'mAb-4', 'mAb-5'])
plt.show()

